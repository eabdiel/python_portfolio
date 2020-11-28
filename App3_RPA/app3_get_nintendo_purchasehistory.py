# Personal project - RPA script that visits the nintendo website, enters the purchase history page, and parses all game & prices into an excel file page by page

import time as t

import openpyxl
import rpa as r

wb = openpyxl.Workbook()
sheet = wb.active
sheet.cell(row=1, column=1).value = 'Name'
sheet.cell(row=1, column=2).value = 'Price'

r.init()

# get URL of old groceries order to re-add all items to cart
# r.url(r.ask('Paste the URL of Purchase History Link and login to Nintendo'))
# r.url('https://ec.nintendo.com/my/#/transactions/1')
r.url('https://ec.nintendo.com/my/#/')
t.sleep(7)

# click on purchase history if logged in
r.click('/html/body/div[1]/div[2]/section/div[1]/section[2]/section/ul/li[3]/a/div[1]')
t.sleep(5)
r.timeout(300)
# set a maximum 5-minute timeout for user to login
# t.sleep(300)

# use exist() function with XPath to check if logged in
if not r.exist('/html/body/div[1]/div[2]/section/div[1]/section[2]/div/div/section[1]/div[2]/div/div[1]/div[2]'):
    r.dom('alert("Purchase History Page not detected after 5 minutes. Bye!")')

# then click on last item arrow
r.click('/html/body/div[1]/div[2]/section/div[1]/section[2]/div/div/ni-pager/section/div/button[9]')

# field using css selector
# key_name1 = 'section.o_c-card-history:nth-child('
# key_name2 = ') > div:nth-child(2) > div:nth-child(1) > div:nth-child(1) > div:nth-child(2) > span:nth-child(1)'
# field using xpath
key_name1 = '/html/body/div[1]/div[2]/section/div[1]/section[2]/div/div/section['
key_name2 = ']/div[2]/div/div[1]/div[2]'

price_part1 = '/html/body/div[1]/div[2]/section/div[1]/section[2]/div/div/section['
price_part2 = ']/div[2]/div/div[2]/div[2]'

purchase_history_link = str(r.url())
purchase_history = purchase_history_link[:42]
purchase_history_number = purchase_history_link[-2:]

# change back to default of 10 seconds to quit fast on error
t.sleep(5)

links = int(purchase_history_number)
row = 1
for link in range(1, links + 1):
    t.sleep(1)
    total_items = int(r.count('section.o_c-card-history')) + 1  # total of items per page
    item = 1
    t.sleep(1)
    while item < total_items:
        # key_full = key_name1 + str(item) + key_name2
        # this f" is for new python 3.6 text format; it allows for changable variables inside the string by using {}
        key_full = f"/html/body/div[1]/div[2]/section/div[1]/section[2]/div/div/section[{str(item)}]" \
                   f"/div[2]/div/div[1]/div[2]"
        price_full = price_part1 + str(item) + price_part2
        var = r.read(key_full)
        t.sleep(1)
        r.timeout(5)
        if not r.exist(price_full):
            var2 = '0.00'
        else:
            var2 = r.read(price_full)
        t.sleep(1)
        sheet.cell(row=row, column=1).value = var
        sheet.cell(row=row, column=2).value = var2
        # you can add records to a list, then use append as well
        # sheet.append(list)
        print(var)
        row += 1
        item += 1
    minus_link = int(purchase_history_number) - link
    purchase_history_link = purchase_history + str(minus_link)
    r.url(purchase_history_link)
r.close()

wb.save('purchase_history.xlsx')
